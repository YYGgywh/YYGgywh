# backend/app/api/admin.py 2026-03-01 10:00:00
# 功能：后台管理API路由

import time
import csv
from io import BytesIO, StringIO
from fastapi import APIRouter, Depends, HTTPException, status, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_
from typing import Optional, List
from app.utils.dependencies import get_db, get_current_admin_user, get_super_admin
from app.models.user import User
from app.models.pan_record import PanRecord
from app.models.comment import Comment
from app.models.system_log import SystemLog
from app.models.system_config import SystemConfig
from app.utils.password import verify_password, hash_password
from app.utils.token import create_access_token
from app.utils.response_formatter import create_success_response, create_error_response
from app.utils.logger import log_system_action

# 尝试导入openpyxl，如果失败则使用CSV作为备选
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter()


# ==================== 管理员登录相关 ====================

class AdminLoginRequest(BaseModel):
    username: str = Field(..., description="手机号或登录名")
    password: str = Field(..., description="密码")


@router.post("/login")
async def admin_login(
    request: AdminLoginRequest,
    db: Session = Depends(get_db)
):
    """管理员登录"""
    user = db.query(User).filter(
        or_(
            User.phone == request.username,
            User.login_name == request.username
        ),
        User.deleted_at.is_(None)
    ).first()
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="用户名或密码错误"
        )
    
    if user.role < 1:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="权限不足，非管理员账户"
        )
    
    # 检查是否设置了后台密码，未设置时使用前台密码
    if not user.admin_password:
        if not user.password:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="账户未设置密码"
            )
        if not verify_password(request.password, user.password):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="用户名或密码错误"
            )
    else:
        # 使用后台密码验证
        if not verify_password(request.password, user.admin_password):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="用户名或密码错误"
            )
    
    if user.status != 1:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="账户已被禁用"
        )
    
    user.last_login_time = int(time.time())
    user.login_count = (user.login_count or 0) + 1
    db.commit()
    
    token_data = {
        "user_id": user.id,
        "phone": user.phone,
        "role": user.role
    }
    token = create_access_token(token_data, is_admin=True)
    
    # 暂时禁用日志记录，排查问题
    # log_system_action(
    #     db=db,
    #     user_id=user.id,
    #     action='login',
    #     module='system',
    #     detail='管理员登录成功'
    # )
    
    return create_success_response({
        "token": token,
        "user": {
            "id": user.id,
            "phone": user.phone,
            "login_name": user.login_name,
            "nickname": user.nickname,
            "avatar": user.avatar,
            "role": user.role
        }
    }, "登录成功")


@router.get("/statistics")
async def get_statistics(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取系统统计数据"""
    user_count = db.query(User).filter(User.deleted_at.is_(None)).count()
    pan_record_count = db.query(PanRecord).count()
    comment_count = db.query(Comment).count()
    pending_audit_count = db.query(PanRecord).filter(PanRecord.audit_status == 0).count()
    
    return create_success_response({
        "users": user_count,
        "pan_records": pan_record_count,
        "comments": comment_count,
        "pending_audit": pending_audit_count
    })


@router.post("/logout")
async def admin_logout(
    current_user: User = Depends(get_current_admin_user)
):
    """管理员登出（前端只需清除token即可）"""
    return create_success_response(None, "登出成功")


# ==================== 用户管理 ====================

class UserListQuery(BaseModel):
    page: int = 1
    page_size: int = 20
    keyword: Optional[str] = None
    status: Optional[int] = None
    role: Optional[int] = None


@router.get("/users")
async def get_user_list(
    query: UserListQuery = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取用户列表"""
    query_obj = db.query(User).filter(User.deleted_at.is_(None))
    
    if query.keyword:
        query_obj = query_obj.filter(
            or_(
                User.phone.contains(query.keyword),
                User.nickname.contains(query.keyword),
                User.login_name.contains(query.keyword)
            )
        )
    
    if query.status is not None:
        query_obj = query_obj.filter(User.status == query.status)
    
    if query.role is not None:
        query_obj = query_obj.filter(User.role == query.role)
    
    total = query_obj.count()
    offset = (query.page - 1) * query.page_size
    users = query_obj.order_by(User.create_time.desc()).offset(offset).limit(query.page_size).all()
    
    user_list = []
    for user in users:
        user_list.append({
            "id": user.id,
            "phone": user.phone,
            "login_name": user.login_name,
            "nickname": user.nickname,
            "avatar": user.avatar,
            "role": user.role,
            "status": user.status,
            "create_time": user.create_time,
            "last_login_time": user.last_login_time,
            "login_count": user.login_count
        })
    
    return create_success_response({
        "list": user_list,
        "total": total,
        "page": query.page,
        "page_size": query.page_size
    })


class DeletedUserListQuery(BaseModel):
    page: int = 1
    page_size: int = 20
    keyword: Optional[str] = None


@router.get("/users/deleted")
async def get_deleted_user_list(
    query: DeletedUserListQuery = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取已删除用户列表"""
    query_obj = db.query(User).filter(User.deleted_at.isnot(None))
    
    if query.keyword:
        query_obj = query_obj.filter(
            or_(
                User.phone.contains(query.keyword),
                User.nickname.contains(query.keyword),
                User.login_name.contains(query.keyword)
            )
        )
    
    total = query_obj.count()
    offset = (query.page - 1) * query.page_size
    users = query_obj.order_by(User.deleted_at.desc()).offset(offset).limit(query.page_size).all()
    
    user_list = []
    for user in users:
        user_list.append({
            "id": user.id,
            "phone": user.phone,
            "login_name": user.login_name,
            "nickname": user.nickname,
            "avatar": user.avatar,
            "role": user.role,
            "status": user.status,
            "create_time": user.create_time,
            "deleted_at": user.deleted_at,
            "last_login_time": user.last_login_time,
            "login_count": user.login_count
        })
    
    return create_success_response({
        "list": user_list,
        "total": total,
        "page": query.page,
        "page_size": query.page_size
    })


@router.get("/users/{user_id}")
async def get_user_detail(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取用户详情"""
    user = db.query(User).filter(
        User.id == user_id,
        User.deleted_at.is_(None)
    ).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 权限检查：普通管理员不能查看超级管理员
    if current_user.role != 99 and user.role == 99:
        raise HTTPException(status_code=403, detail="无权查看超级管理员信息")
    
    return create_success_response({
        "id": user.id,
        "phone": user.phone,
        "login_name": user.login_name,
        "email": user.email,
        "nickname": user.nickname,
        "avatar": user.avatar,
        "wechat_openid": user.wechat_openid,
        "gender": user.gender,
        "birth_calendar_type": user.birth_calendar_type,
        "birth_year": user.birth_year,
        "birth_month": user.birth_month,
        "birth_day": user.birth_day,
        "birth_hour": user.birth_hour,
        "birth_minute": user.birth_minute,
        "birth_second": user.birth_second,
        "login_name_modify_count": user.login_name_modify_count,
        "login_name_modify_time": user.login_name_modify_time,
        "is_old_user": user.is_old_user,
        "ext_info": user.ext_info,
        "role": user.role,
        "status": user.status,
        "is_active": user.is_active,
        "create_time": user.create_time,
        "update_time": user.update_time,
        "deleted_at": user.deleted_at,
        "last_login_time": user.last_login_time,
        "last_login_ip": user.last_login_ip,
        "login_count": user.login_count,
        "has_admin_password": user.admin_password is not None  # 是否设置了后台密码
    })


class UpdateUserRequest(BaseModel):
    nickname: Optional[str] = None
    role: Optional[int] = None
    status: Optional[int] = None
    login_name: Optional[str] = None
    email: Optional[str] = None
    gender: Optional[int] = None


class UpdateUserPasswordRequest(BaseModel):
    new_password: str = Field(..., min_length=6, description="新密码，至少6位")


@router.put("/users/{user_id}")
async def update_user(
    user_id: int,
    request: UpdateUserRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """更新用户信息"""
    user = db.query(User).filter(
        User.id == user_id,
        User.deleted_at.is_(None)
    ).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 权限检查：普通管理员只能编辑普通用户（role=0）
    if current_user.role != 99 and user.role >= 1:
        raise HTTPException(status_code=403, detail="无权编辑管理员或超级管理员信息")
    
    if request.nickname is not None:
        user.nickname = request.nickname
    if request.role is not None:
        if current_user.role != 99 and request.role == 99:
            raise HTTPException(status_code=403, detail="无权设置为超级管理员")
        user.role = request.role
    if request.status is not None:
        user.status = request.status
    if request.login_name is not None:
        user.login_name = request.login_name
    if request.email is not None:
        user.email = request.email
    if request.gender is not None:
        user.gender = request.gender
    
    user.update_time = int(time.time())
    db.commit()
    
    return create_success_response(None, "更新成功")


@router.put("/users/{user_id}/password")
async def update_user_password(
    user_id: int,
    request: UpdateUserPasswordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """修改用户密码"""
    user = db.query(User).filter(
        User.id == user_id,
        User.deleted_at.is_(None)
    ).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 权限检查：普通管理员只能修改普通用户（role=0）的密码，超级管理员可以修改所有用户
    if current_user.role != 99 and user.role >= 1:
        raise HTTPException(status_code=403, detail="无权修改管理员或超级管理员的密码")
    
    # 检查密码强度
    from app.utils.password import check_password_strength
    password_strength = check_password_strength(request.new_password)
    if password_strength['level'] in ['极弱']:
        raise HTTPException(status_code=400, detail=f"密码强度不足：{password_strength['message']}")
    
    # 更新密码
    user.password = hash_password(request.new_password)
    user.update_time = int(time.time())
    db.commit()
    
    log_system_action(
        db=db,
        user_id=current_user.id,
        action='update_password',
        module='user',
        target_type='user',
        target_id=user_id,
        detail=f"修改用户 ID: {user_id}, 手机号: {user.phone} 的前台密码"
    )
    
    return create_success_response(None, "密码修改成功")


class UpdateAdminPasswordRequest(BaseModel):
    new_password: str = Field(..., min_length=6, description="新的后台密码，至少6位")


@router.put("/users/{user_id}/admin-password")
async def update_admin_password(
    user_id: int,
    request: UpdateAdminPasswordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """修改用户后台密码"""
    user = db.query(User).filter(
        User.id == user_id,
        User.deleted_at.is_(None)
    ).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 权限检查：普通管理员只能修改普通用户的密码，超级管理员可以修改所有用户
    if current_user.role != 99 and user.role >= 1:
        raise HTTPException(status_code=403, detail="无权修改管理员或超级管理员的后台密码")
    
    # 检查密码强度
    from app.utils.password import check_password_strength
    password_strength = check_password_strength(request.new_password)
    if password_strength['level'] in ['极弱']:
        raise HTTPException(status_code=400, detail=f"密码强度不足：{password_strength['message']}")
    
    # 更新后台密码
    user.admin_password = hash_password(request.new_password)
    user.update_time = int(time.time())
    db.commit()
    
    log_system_action(
        db=db,
        user_id=current_user.id,
        action='update_admin_password',
        module='user',
        target_type='user',
        target_id=user_id,
        detail=f"修改用户 ID: {user_id}, 手机号: {user.phone} 的后台密码"
    )
    
    return create_success_response(None, "后台密码修改成功")


@router.delete("/users/{user_id}")
async def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """删除用户（软删除）"""
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="不能删除自己")
    
    user = db.query(User).filter(
        User.id == user_id,
        User.deleted_at.is_(None)
    ).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 权限检查：普通管理员只能删除普通用户（role=0），超级管理员可以删除所有人（除了自己）
    if current_user.role != 99 and user.role >= 1:
        raise HTTPException(status_code=403, detail="无权删除管理员或超级管理员")
    
    user.deleted_at = int(time.time())
    db.commit()
    
    log_system_action(
        db=db,
        user_id=current_user.id,
        action='delete',
        module='user',
        target_type='user',
        target_id=user_id,
        detail=f"删除用户 ID: {user_id}, 手机号: {user.phone}"
    )
    
    return create_success_response(None, "删除成功")


@router.put("/users/{user_id}/restore")
async def restore_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """恢复已删除用户"""
    user = db.query(User).filter(
        User.id == user_id,
        User.deleted_at.isnot(None)
    ).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在或未被删除")
    
    # 权限检查
    if user.role >= current_user.role and current_user.role != 99:
        raise HTTPException(status_code=403, detail="无权恢复权限相等或更高的用户")
    
    user.deleted_at = None
    user.update_time = int(time.time())
    db.commit()
    
    log_system_action(
        db=db,
        user_id=current_user.id,
        action='restore',
        module='user',
        target_type='user',
        target_id=user_id,
        detail=f"恢复用户 ID: {user_id}, 手机号: {user.phone}"
    )
    
    return create_success_response(None, "恢复成功")


@router.delete("/users/{user_id}/permanent")
async def permanent_delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_super_admin)
):
    """永久删除用户（真正从数据库删除）"""
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 防止删除自己
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="不能永久删除自己")
    
    # 只有超级管理员才能永久删除
    if current_user.role != 99:
        raise HTTPException(status_code=403, detail="只有超级管理员才能永久删除用户")
    
    db.delete(user)
    db.commit()
    
    log_system_action(
        db=db,
        user_id=current_user.id,
        action='permanent_delete',
        module='user',
        target_type='user',
        target_id=user_id,
        detail=f"永久删除用户 ID: {user_id}, 手机号: {user.phone}"
    )
    
    return create_success_response(None, "永久删除成功")


# ==================== 排盘记录管理 ====================

class PanRecordListQuery(BaseModel):
    page: int = 1
    page_size: int = 20
    audit_status: Optional[int] = None
    user_id: Optional[int] = None


@router.get("/pan-records")
async def get_pan_record_list(
    query: PanRecordListQuery = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取排盘记录列表"""
    query_obj = db.query(PanRecord).filter(PanRecord.deleted_at.is_(None))
    
    if query.audit_status is not None:
        query_obj = query_obj.filter(PanRecord.audit_status == query.audit_status)
    
    if query.user_id:
        query_obj = query_obj.filter(PanRecord.user_id == query.user_id)
    
    total = query_obj.count()
    offset = (query.page - 1) * query.page_size
    records = query_obj.order_by(PanRecord.create_time.desc()).offset(offset).limit(query.page_size).all()
    
    record_list = []
    for record in records:
        record_list.append({
            "id": record.id,
            "user_id": record.user_id,
            "pan_type": record.pan_type,
            "audit_status": record.audit_status,
            "is_visible": record.is_visible,
            "create_time": record.create_time,
            "audit_time": record.audit_time
        })
    
    return create_success_response({
        "list": record_list,
        "total": total,
        "page": query.page,
        "page_size": query.page_size
    })


@router.get("/pan-records/{record_id}")
async def get_pan_record_detail(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取排盘记录详情"""
    record = db.query(PanRecord).filter(
        PanRecord.id == record_id,
        PanRecord.deleted_at.is_(None)
    ).first()
    
    if not record:
        raise HTTPException(status_code=404, detail="排盘记录不存在")
    
    return create_success_response({
        "id": record.id,
        "user_id": record.user_id,
        "pan_type": record.pan_type,
        "pan_params": record.pan_params,
        "pan_result": record.pan_result,
        "audit_status": record.audit_status,
        "audit_remark": record.audit_remark,
        "audit_time": record.audit_time,
        "audit_user_id": record.audit_user_id,
        "is_visible": record.is_visible,
        "supplement": record.supplement,
        "create_time": record.create_time,
        "update_time": record.update_time,
        "deleted_at": record.deleted_at
    })


class AuditPanRecordRequest(BaseModel):
    audit_status: int
    audit_remark: Optional[str] = None


@router.put("/pan-records/{record_id}/audit")
async def audit_pan_record(
    record_id: int,
    request: AuditPanRecordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """审核排盘记录"""
    record = db.query(PanRecord).filter(
        PanRecord.id == record_id,
        PanRecord.deleted_at.is_(None)
    ).first()
    
    if not record:
        raise HTTPException(status_code=404, detail="排盘记录不存在")
    
    record.audit_status = request.audit_status
    record.audit_remark = request.audit_remark
    record.audit_time = int(time.time())
    record.audit_user_id = current_user.id
    db.commit()
    
    return create_success_response(None, "审核成功")


@router.delete("/pan-records/{record_id}")
async def delete_pan_record(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """删除排盘记录（软删除）"""
    record = db.query(PanRecord).filter(
        PanRecord.id == record_id,
        PanRecord.deleted_at.is_(None)
    ).first()
    
    if not record:
        raise HTTPException(status_code=404, detail="排盘记录不存在")
    
    record.deleted_at = int(time.time())
    db.commit()
    
    return create_success_response(None, "删除成功")


# ==================== 批量操作 ====================

class BatchDeleteRequest(BaseModel):
    record_ids: List[int] = Field(..., description="要删除的记录ID列表")


class BatchAuditRequest(BaseModel):
    record_ids: List[int] = Field(..., description="要审核的记录ID列表")
    audit_status: int = Field(..., description="审核状态：1-通过，2-拒绝")
    audit_remark: Optional[str] = Field(None, description="审核备注")


@router.post("/pan-records/batch-delete")
async def batch_delete_pan_records(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """批量删除排盘记录（软删除）"""
    if not request.record_ids:
        raise HTTPException(status_code=400, detail="记录ID列表不能为空")
    
    # 查询所有有效的记录
    records = db.query(PanRecord).filter(
        PanRecord.id.in_(request.record_ids),
        PanRecord.deleted_at.is_(None)
    ).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="未找到有效的排盘记录")
    
    current_time = int(time.time())
    deleted_count = 0
    
    for record in records:
        record.deleted_at = current_time
        deleted_count += 1
    
    db.commit()
    
    return create_success_response(
        {"deleted_count": deleted_count},
        f"成功删除 {deleted_count} 条记录"
    )


@router.post("/pan-records/batch-audit")
async def batch_audit_pan_records(
    request: BatchAuditRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """批量审核排盘记录"""
    if not request.record_ids:
        raise HTTPException(status_code=400, detail="记录ID列表不能为空")
    
    if request.audit_status not in [1, 2]:
        raise HTTPException(status_code=400, detail="审核状态必须是1（通过）或2（拒绝）")
    
    # 查询所有待审核的记录
    records = db.query(PanRecord).filter(
        PanRecord.id.in_(request.record_ids),
        PanRecord.deleted_at.is_(None),
        PanRecord.audit_status == 0
    ).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="未找到待审核的排盘记录")
    
    current_time = int(time.time())
    audited_count = 0
    
    for record in records:
        record.audit_status = request.audit_status
        record.audit_remark = request.audit_remark
        record.audit_time = current_time
        record.audit_user_id = current_user.id
        audited_count += 1
    
    db.commit()
    
    status_text = "通过" if request.audit_status == 1 else "拒绝"
    return create_success_response(
        {"audited_count": audited_count},
        f"成功审核 {audited_count} 条记录，状态：{status_text}"
    )


# ==================== 已删除记录管理 ====================

class DeletedPanRecordListQuery(BaseModel):
    page: int = 1
    page_size: int = 20
    keyword: Optional[str] = None


@router.get("/pan-records/deleted")
async def get_deleted_pan_record_list(
    query: DeletedPanRecordListQuery = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取已删除的排盘记录列表"""
    query_obj = db.query(PanRecord).filter(PanRecord.deleted_at.isnot(None))
    
    if query.keyword:
        query_obj = query_obj.filter(
            or_(
                PanRecord.user_id == query.keyword,
                PanRecord.pan_type.ilike(f"%{query.keyword}%")
            )
        )
    
    total = query_obj.count()
    offset = (query.page - 1) * query.page_size
    records = query_obj.order_by(PanRecord.deleted_at.desc()).offset(offset).limit(query.page_size).all()
    
    result = []
    for record in records:
        result.append({
            "id": record.id,
            "user_id": record.user_id,
            "pan_type": record.pan_type,
            "audit_status": record.audit_status,
            "supplement": record.supplement,
            "create_time": record.create_time,
            "update_time": record.update_time,
            "deleted_at": record.deleted_at
        })
    
    return create_success_response({
        "list": result,
        "total": total,
        "page": query.page,
        "page_size": query.page_size
    })


@router.put("/pan-records/{record_id}/restore")
async def restore_pan_record(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """恢复已删除的排盘记录"""
    record = db.query(PanRecord).filter(
        PanRecord.id == record_id,
        PanRecord.deleted_at.isnot(None)
    ).first()
    
    if not record:
        raise HTTPException(status_code=404, detail="排盘记录不存在")
    
    record.deleted_at = None
    record.update_time = int(time.time())
    db.commit()
    
    return create_success_response(None, "恢复成功")


@router.delete("/pan-records/{record_id}/permanent")
async def permanent_delete_pan_record(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """永久删除排盘记录"""
    record = db.query(PanRecord).filter(
        PanRecord.id == record_id,
        PanRecord.deleted_at.isnot(None)
    ).first()
    
    if not record:
        raise HTTPException(status_code=404, detail="排盘记录不存在")
    
    db.delete(record)
    db.commit()
    
    return create_success_response(None, "永久删除成功")


@router.post("/pan-records/batch-permanent-delete")
async def batch_permanent_delete_pan_records(
    request: BatchDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """批量永久删除排盘记录"""
    if not request.record_ids:
        raise HTTPException(status_code=400, detail="记录ID列表不能为空")
    
    # 查询所有已删除的记录
    records = db.query(PanRecord).filter(
        PanRecord.id.in_(request.record_ids),
        PanRecord.deleted_at.isnot(None)
    ).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="未找到已删除的排盘记录")
    
    deleted_count = 0
    
    for record in records:
        db.delete(record)
        deleted_count += 1
    
    db.commit()
    
    return create_success_response(
        {"deleted_count": deleted_count},
        f"成功永久删除 {deleted_count} 条记录"
    )


# ==================== 数据导出 ====================

class ExportRequest(BaseModel):
    format: str = Field(..., description="导出格式：excel 或 csv")
    filters: Optional[dict] = Field(None, description="筛选条件")
    record_ids: Optional[List[int]] = Field(None, description="要导出的记录ID列表，为空则导出所有符合条件的记录")


@router.post("/pan-records/export")
async def export_pan_records(
    request: ExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """导出排盘记录"""
    # 构建查询
    query_obj = db.query(PanRecord).filter(PanRecord.deleted_at.is_(None))
    
    # 优先按记录ID导出
    if request.record_ids:
        query_obj = query_obj.filter(PanRecord.id.in_(request.record_ids))
    else:
        # 应用筛选条件
        if request.filters:
            if request.filters.get("keyword"):
                query_obj = query_obj.filter(
                    or_(
                        PanRecord.user_id == request.filters["keyword"],
                        PanRecord.pan_type.ilike(f"%{request.filters['keyword']}%")
                    )
                )
            if request.filters.get("audit_status") is not None:
                query_obj = query_obj.filter(
                    PanRecord.audit_status == request.filters["audit_status"]
                )
            if request.filters.get("pan_type"):
                query_obj = query_obj.filter(
                    PanRecord.pan_type == request.filters["pan_type"]
                )
    
    # 获取所有记录
    records = query_obj.all()
    
    if not records:
        raise HTTPException(status_code=404, detail="没有符合条件的记录")
    
    if request.format == "excel" and HAS_OPENPYXL:
        return export_to_excel(records)
    else:
        return export_to_csv(records)


def export_to_excel(records):
    """导出为Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "排盘记录"
    
    # 设置表头
    headers = ["ID", "用户ID", "排盘类型", "审核状态", "创建时间", "更新时间", "审核时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 填充数据
    for row, record in enumerate(records, 2):
        status_map = {0: "待审核", 1: "已通过", 2: "已拒绝"}
        pan_type_map = {"liuyao": "六爻", "bazi": "八字", "ziwei": "紫微斗数", "qimen": "奇门遁甲"}
        
        ws.cell(row=row, column=1, value=record.id)
        ws.cell(row=row, column=2, value=record.user_id)
        ws.cell(row=row, column=3, value=pan_type_map.get(record.pan_type, record.pan_type))
        ws.cell(row=row, column=4, value=status_map.get(record.audit_status, "未知"))
        ws.cell(row=row, column=5, value=time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(record.create_time)) if record.create_time else "")
        ws.cell(row=row, column=6, value=time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(record.update_time)) if record.update_time else "")
        ws.cell(row=row, column=7, value=time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(record.audit_time)) if record.audit_time else "")
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # 保存到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    import urllib.parse
    filename = f"排盘记录_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={encoded_filename}; filename*=UTF-8''{encoded_filename}"
        }
    )


def export_to_csv(records):
    """导出为CSV文件"""
    output = StringIO()
    writer = csv.writer(output)
    
    # 写入表头
    writer.writerow(["ID", "用户ID", "排盘类型", "审核状态", "创建时间", "更新时间", "审核时间"])
    
    # 写入数据
    for record in records:
        status_map = {0: "待审核", 1: "已通过", 2: "已拒绝"}
        pan_type_map = {"liuyao": "六爻", "bazi": "八字", "ziwei": "紫微斗数", "qimen": "奇门遁甲"}
        
        writer.writerow([
            record.id,
            record.user_id,
            pan_type_map.get(record.pan_type, record.pan_type),
            status_map.get(record.audit_status, "未知"),
            time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(record.create_time)) if record.create_time else "",
            time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(record.update_time)) if record.update_time else "",
            time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(record.audit_time)) if record.audit_time else ""
        ])
    
    output.seek(0)
    import urllib.parse
    filename = f"排盘记录_{time.strftime('%Y%m%d_%H%M%S')}.csv"
    encoded_filename = urllib.parse.quote(filename)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={encoded_filename}; filename*=UTF-8''{encoded_filename}"
        }
    )


# ==================== 评论管理 ====================

class CommentListQuery(BaseModel):
    page: int = 1
    page_size: int = 20
    audit_status: Optional[int] = None
    user_id: Optional[int] = None


@router.get("/comments")
async def get_comment_list(
    query: CommentListQuery = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取评论列表"""
    query_obj = db.query(Comment).filter(Comment.deleted_at.is_(None))
    
    if query.audit_status is not None:
        query_obj = query_obj.filter(Comment.audit_status == query.audit_status)
    
    if query.user_id:
        query_obj = query_obj.filter(Comment.user_id == query.user_id)
    
    total = query_obj.count()
    offset = (query.page - 1) * query.page_size
    comments = query_obj.order_by(Comment.create_time.desc()).offset(offset).limit(query.page_size).all()
    
    comment_list = []
    for comment in comments:
        comment_list.append({
            "id": comment.id,
            "pan_record_id": comment.pan_record_id,
            "user_id": comment.user_id,
            "content": comment.content,
            "audit_status": comment.audit_status,
            "is_visible": comment.is_visible,
            "create_time": comment.create_time,
            "audit_time": comment.audit_time
        })
    
    return create_success_response({
        "list": comment_list,
        "total": total,
        "page": query.page,
        "page_size": query.page_size
    })


class AuditCommentRequest(BaseModel):
    audit_status: int
    audit_remark: Optional[str] = None


@router.put("/comments/{comment_id}/audit")
async def audit_comment(
    comment_id: int,
    request: AuditCommentRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """审核评论"""
    comment = db.query(Comment).filter(
        Comment.id == comment_id,
        Comment.deleted_at.is_(None)
    ).first()
    
    if not comment:
        raise HTTPException(status_code=404, detail="评论不存在")
    
    comment.audit_status = request.audit_status
    comment.audit_remark = request.audit_remark
    comment.audit_time = int(time.time())
    comment.audit_user_id = current_user.id
    db.commit()
    
    return create_success_response(None, "审核成功")


@router.delete("/comments/{comment_id}")
async def delete_comment(
    comment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """删除评论（软删除）"""
    comment = db.query(Comment).filter(
        Comment.id == comment_id,
        Comment.deleted_at.is_(None)
    ).first()
    
    if not comment:
        raise HTTPException(status_code=404, detail="评论不存在")
    
    comment.deleted_at = int(time.time())
    db.commit()
    
    return create_success_response(None, "删除成功")


# ==================== 系统日志 ====================

class SystemLogListQuery(BaseModel):
    page: int = 1
    page_size: int = 20
    module: Optional[str] = None
    action: Optional[str] = None
    user_id: Optional[int] = None


@router.get("/system-logs")
async def get_system_log_list(
    query: SystemLogListQuery = Depends(),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取系统日志列表"""
    query_obj = db.query(SystemLog)
    
    if query.module:
        query_obj = query_obj.filter(SystemLog.module == query.module)
    
    if query.action:
        query_obj = query_obj.filter(SystemLog.action == query.action)
    
    if query.user_id:
        query_obj = query_obj.filter(SystemLog.user_id == query.user_id)
    
    total = query_obj.count()
    offset = (query.page - 1) * query.page_size
    logs = query_obj.order_by(SystemLog.create_time.desc()).offset(offset).limit(query.page_size).all()
    
    log_list = []
    for log in logs:
        log_list.append({
            "id": log.id,
            "user_id": log.user_id,
            "action": log.action,
            "module": log.module,
            "target_type": log.target_type,
            "target_id": log.target_id,
            "detail": log.detail,
            "ip": log.ip,
            "create_time": log.create_time
        })
    
    return create_success_response({
        "list": log_list,
        "total": total,
        "page": query.page,
        "page_size": query.page_size
    })


@router.get("/system-logs/export")
async def export_system_logs(
    module: Optional[str] = None,
    action: Optional[str] = None,
    user_id: Optional[int] = None,
    start_time: Optional[int] = None,
    end_time: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """导出系统日志"""
    import csv
    from io import StringIO
    from fastapi import Response
    
    query_obj = db.query(SystemLog)
    
    if module:
        query_obj = query_obj.filter(SystemLog.module == module)
    
    if action:
        query_obj = query_obj.filter(SystemLog.action == action)
    
    if user_id:
        query_obj = query_obj.filter(SystemLog.user_id == user_id)
    
    if start_time:
        query_obj = query_obj.filter(SystemLog.create_time >= start_time)
    
    if end_time:
        query_obj = query_obj.filter(SystemLog.create_time <= end_time)
    
    logs = query_obj.order_by(SystemLog.create_time.desc()).all()
    
    # 创建CSV文件
    output = StringIO()
    writer = csv.writer(output)
    
    # 写入表头
    writer.writerow(["ID", "用户ID", "操作", "模块", "目标类型", "目标ID", "详细信息", "IP地址", "操作时间"])
    
    # 写入数据
    for log in logs:
        writer.writerow([
            log.id,
            log.user_id,
            log.action,
            log.module,
            log.target_type or "",
            log.target_id or "",
            log.detail or "",
            log.ip or "",
            log.create_time
        ])
    
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=system_logs_{int(time.time())}.csv"
        }
    )


# ==================== 系统配置 ====================

@router.get("/system-configs")
async def get_system_configs(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """获取系统配置列表"""
    configs = db.query(SystemConfig).all()
    
    config_list = []
    for config in configs:
        config_list.append({
            "id": config.id,
            "key": config.key,
            "value": config.value,
            "description": config.description,
            "update_time": config.update_time
        })
    
    return create_success_response(config_list)


class UpdateSystemConfigRequest(BaseModel):
    value: str


@router.put("/system-configs/{config_key}")
async def update_system_config(
    config_key: str,
    request: UpdateSystemConfigRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_super_admin)
):
    """更新系统配置"""
    config = db.query(SystemConfig).filter(SystemConfig.key == config_key).first()
    
    if not config:
        raise HTTPException(status_code=404, detail="配置项不存在")
    
    config.value = request.value
    config.update_time = int(time.time())
    db.commit()
    
    return create_success_response(None, "更新成功")
